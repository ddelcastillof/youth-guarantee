"""
Health intervention only.

Adds an effect size to reg_health_wellbeing.xlsx that applies to those aged
under 25, so the intervention is simulated for that age range alone.

Called by src/00_stage_scenario.py once the pristine inputs are in place; it
edits the workbook where it stands and does not manage backups itself.
"""

from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = "reg_health_wellbeing.xlsx"
SHEET = "DHE_MCS1"
REGRESSOR = "AgeUnder25"
COEFFICIENT = 0.529
VARIANCE = 1e-10


def apply(simpaths_input: Path) -> None:
    working = simpaths_input / WORKBOOK

    workbook = load_workbook(working)
    sheet = workbook[SHEET]

    regressors = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1)]
    if REGRESSOR in regressors:
        # Staging restores the pristine workbook first, so this can only mean the
        # pristine snapshot itself already carries the intervention
        raise RuntimeError(
            f"{REGRESSOR} is already present in {SHEET} of the restored {WORKBOOK}. "
            "The pristine snapshot is contaminated; re-create it from a clean SimPaths input."
        )

    covariances = sheet.max_column - 2
    sheet.append([REGRESSOR, COEFFICIENT] + [0.0] * covariances)

    new_column = sheet.max_column + 1
    sheet.cell(row=1, column=new_column, value=REGRESSOR)
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=new_column, value=0.0)

    sheet.cell(row=sheet.max_row, column=new_column, value=VARIANCE)

    workbook.save(working)
    print(f"Added {REGRESSOR} to {SHEET} of {WORKBOOK}")
