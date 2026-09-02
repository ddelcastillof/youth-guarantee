"""
Adding the health intervention effect for those between 18 and 24 (under 25)
This script will add an effect size to simulate the health intervention only 
in the reg_health_wellbeing.xlsx that will be only applied to those in the specific age range.
The script will only run under scenario 'hi_only'
"""

import os
import sys
from pathlib import Path
from shutil import copy

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent

SIMPATHS_REPO = REPO_ROOT.parent / "SimPaths" / "input"
reg_baseline = REPO_ROOT / "data" / "simpaths_output" / "baseline"
reg_hi_only = REPO_ROOT / "data" / "simpaths_output" / "hi-only"

WORKBOOK = "reg_health_wellbeing.xlsx"
SHEET = "DHE_MCS1"
REGRESSOR = "AgeUnder25"
COEFFICIENT = 0.529
VARIANCE = 1e-10

scenario = os.environ.get("SCENARIO")
if not scenario:
    sys.exit("SCENARIO env var not set")

if scenario.replace("_", "-") != "hi-only":
    sys.exit("scenario is not health intervention only")

print("scenario is health intervention only. Adding effect on target population")

source = Path(SIMPATHS_REPO, WORKBOOK)
backup = Path(reg_baseline, WORKBOOK)
working = Path(reg_hi_only, WORKBOOK)

reg_baseline.mkdir(parents=True, exist_ok=True)
reg_hi_only.mkdir(parents=True, exist_ok=True)

# The backup is captured once and reused afterwards, so a rerun cannot overwrite
# the unmodified coefficients with an input file that already carries the effect.
if not backup.is_file():
    copy(source, backup)
copy(backup, working)

workbook = load_workbook(working)
sheet = workbook[SHEET]

regressors = [row[0].value for row in sheet.iter_rows(min_row=2, max_col=1)]
if REGRESSOR in regressors:
    sys.exit(f"{REGRESSOR} is already present in {SHEET}")

# DHE_MCS1 holds a variance-covariance matrix: REGRESSOR, COEFFICIENT, then one
# column per regressor. A new regressor therefore needs a row and a column.
covariances = sheet.max_column - 2
sheet.append([REGRESSOR, COEFFICIENT] + [0.0] * covariances)

new_column = sheet.max_column + 1
sheet.cell(row=1, column=new_column, value=REGRESSOR)
for row in range(2, sheet.max_row + 1):
    sheet.cell(row=row, column=new_column, value=0.0)

# A zero on the diagonal makes the matrix singular when SimPaths draws from it.
sheet.cell(row=sheet.max_row, column=new_column, value=VARIANCE)

workbook.save(working)
copy(working, source)

print(f"Added {REGRESSOR} to {SHEET} and copied {WORKBOOK} into {SIMPATHS_REPO}")
